from __future__ import annotations

from itertools import chain
import math
from pathlib import Path
from tempfile import NamedTemporaryFile

import duckdb
import plotly.express as px
import plotly.graph_objects as go
import polars as pl
import pyarrow as pa
import streamlit as st
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# App configuration
# ---------------------------------------------------------------------------

# DEFAULT_EXCEL_PATH = Path(r"D:\MARINE REPORTS\REM CASES\2024.xlsx")
DEFAULT_EXCEL_PATH = Path(r"D:\MARINE REPORTS\REM CASES\Remittances10.xlsx")
DEFAULT_SHEET = "Sheet1"
LARGE_WORKBOOK_BYTES = 250 * 1024 * 1024
MAX_STYLED_CELLS = 250_000

# Some exported sheets do not contain a header row. When that happens, the
# first row is data, so we restore the expected remittance schema manually.
REMITTANCE_COLUMNS = [
    "INSTANCE_ID",
    "RMT_BNK",
    "RMT_BCH",
    "RMT_CNM",
    "RMT_AD1",
    "RMT_AD2",
    "RMT_AD3",
    "RMT_IDD",
    "RMT_VAT",
    "RMT_BNM",
    "RMT_BAD",
    "RMT_BCN",
    "RMT_ECA",
    "RMT_FEP",
    "RMT_CUR",
    "RMT_CUX",
    "RMT_AMT",
    "RMT_TOP",
    "RMT_TOD",
    "RMT_TDX",
    "RMT_REF",
    "RMT_LCN",
    "RMT_DAT",
    "RMT_BAL",
    "RMT_OUT",
    "RMT_LCD",
    "RMT_LCE",
    "RMT_DES",
    "RMT_TPX",
    "CNL_DAT",
    "RMT_CUO",
    "RMT_YEA",
    "RMT_SER",
    "RMT_NBR",
    "RMT_COM",
    "RMT_OHF",
    "DEC_REF_YER",
    "IDE_CUO_COD",
    "IDE_REG_SER",
    "IDE_REG_NBR",
    "IDE_REG_DAT",
    "CMP_CON_COD",
    "CMP_CON_NAM",
    "DEC_COD",
    "DEC_NAM",
]

RENAMED_COLUMNS = {
    "TO_CHAR(G.CMP_CON_NAM)": "CMP_CON_NAM",
    "TO_CHAR(G.DEC_NAM)": "DEC_NAM",
}

AMOUNT_COLUMNS = ["RMT_AMT", "RMT_OUT", "RMT_BAL"]
DATE_COLUMNS = ["RMT_DAT", "IDE_REG_DAT", "CNL_DAT"]
TEXT_SEARCH_COLUMNS = ["RMT_CNM", "RMT_BNM", "RMT_DES", "RMT_REF", "DEC_COD", "CMP_CON_COD"]

MONEY_COLUMNS = {
    "RMT_AMT",
    "RMT_OUT",
    "RMT_BAL",
    "total_amount",
    "total_out",
    "total_balance",
    "total_sent_amount",
    "import_linked_amount",
    "no_import_amount",
    "no_cusdec_amount",
    "same_day_amount",
    "high_value_amount",
    "generic_description_amount",
    "max_amount",
    "open_balance",
    "total_open_balance",
    "max_open_balance",
    "avg_amount",
    "avg_sent_amount",
    "avg_balance",
    "avg_open_balance",
    "max_balance",
    "beneficiary_avg_amount",
    "residual_ratio",
}
PERCENT_COLUMNS = {
    "balance Ratio",
    "balance_to_amount_ratio",
    "import_linked_rate",
    "no_import_rate",
    "no_cusdec_rate",
    "open_balance_rate",
}
DATE_DISPLAY_COLUMNS = {"RMT_DAT", "IDE_REG_DAT", "CNL_DAT", "first_rmt_date", "last_rmt_date"}
NUMBER_COLUMNS = {
    "records",
    "transfers",
    "references",
    "ref_count",
    "currency_count",
    "open_balance_transfers",
    "remittance_count",
    "import_linked_records",
    "import_declaration_count",
    "no_import_records",
    "no_cusdec_records",
    "same_day_transfer_count",
    "high_value_records",
    "generic_description_records",
    "remitter_count",
    "beneficiary_count",
    "cnm_count",
    "bnm_count",
    "consignee_count",
    "declarant_count",
    "customs_key_count",
    "relationship_count",
    "risk_score",
    "active_days",
    "days_unwritten",
    "date_gap_days",
    "bnm_no_cusdec_count",
    "cnm_no_cusdec_count",
    "pair_no_cusdec_count",
    "same_day_beneficiary_transfers",
}

MISSING_TEXT_SQL = "('', 'nan', 'none', 'null')"
NO_CUSDEC_CONDITION = f"""
(
    coalesce(lower(trim(cast(IDE_REG_NBR as varchar))), '') in {MISSING_TEXT_SQL}
    or IDE_REG_DAT is null
    or coalesce(lower(trim(cast(CMP_CON_COD as varchar))), '') in {MISSING_TEXT_SQL}
)
"""

DECLARANT_NODE_SQL = """
concat(
    coalesce(cast(DEC_COD as varchar), 'Unknown'),
    ' | ',
    left(replace(coalesce(cast(DEC_NAM as varchar), 'Unknown'), chr(10), ' '), 58)
)
"""

CONSIGNEE_NODE_SQL = """
concat(
    coalesce(cast(CMP_CON_COD as varchar), 'Unknown'),
    ' | ',
    left(replace(coalesce(cast(CMP_CON_NAM as varchar), 'Unknown'), chr(10), ' '), 58)
)
"""


st.set_page_config(page_title="Remittance Review Dashboard", layout="wide")


# ---------------------------------------------------------------------------
# Loading and common helpers
# ---------------------------------------------------------------------------

def clean_column_name(name: object) -> str:
    return str(name).strip().replace("\n", " ").replace("\r", " ").replace('"', "")


def sql_string(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_like(value: str) -> str:
    safe = value.casefold().replace("%", "").replace("_", "")
    return sql_string(f"%{safe}%")


@st.cache_data(show_spinner=False)
def get_sheet_names(path: str) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def load_excel_streaming(path: str, sheet_name: str) -> pl.DataFrame:
    """Read very large xlsx files row-by-row to avoid Calamine memory spikes."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row is None:
            return pl.DataFrame()

        first_values = list(first_row)
        first_as_text = [clean_column_name(value) for value in first_values]
        has_header = "INSTANCE_ID" in first_as_text

        if has_header:
            columns = first_as_text
            data_rows = rows
        elif len(first_values) == len(REMITTANCE_COLUMNS):
            columns = REMITTANCE_COLUMNS
            data_rows = chain([first_values], rows)
        else:
            columns = [f"column_{idx + 1}" for idx in range(len(first_values))]
            data_rows = chain([first_values], rows)

        records = {column: [] for column in columns}
        column_count = len(columns)
        for row in data_rows:
            values = list(row[:column_count])
            if len(values) < column_count:
                values.extend([None] * (column_count - len(values)))
            if not any(value is not None for value in values):
                continue
            for column, value in zip(columns, values):
                records[column].append(value)

        return pl.DataFrame(records, strict=False)
    finally:
        workbook.close()


@st.cache_data(show_spinner="Loading workbook...")
def load_excel(path: str, sheet_name: str) -> pl.DataFrame:
    path_obj = Path(path)
    if path_obj.stat().st_size >= LARGE_WORKBOOK_BYTES:
        df = load_excel_streaming(path, sheet_name)
    else:
        df = pl.read_excel(path, sheet_name=sheet_name, engine="calamine")
        cleaned_columns = [clean_column_name(col) for col in df.columns]
        if "INSTANCE_ID" not in cleaned_columns and df.width == len(REMITTANCE_COLUMNS):
            first_row = dict(zip(REMITTANCE_COLUMNS, df.columns))
            df.columns = REMITTANCE_COLUMNS
            first_df = pl.DataFrame({col: [value] for col, value in first_row.items()}, strict=False)
            df = pl.concat([first_df, df], how="vertical_relaxed")

    df = df.rename({col: clean_column_name(col) for col in df.columns})
    df = df.rename({src: dst for src, dst in RENAMED_COLUMNS.items() if src in df.columns and dst not in df.columns})

    for col in AMOUNT_COLUMNS:
        if col in df.columns:
            df = df.with_columns(pl.col(col).cast(pl.Float64, strict=False))

    for col in DATE_COLUMNS:
        if col in df.columns and df.schema[col] not in (pl.Date, pl.Datetime):
            df = df.with_columns(pl.col(col).cast(pl.Datetime, strict=False))

    return df


def arrow_table(df: pl.DataFrame) -> pa.Table:
    return df.to_arrow()


def run_query(table: pa.Table, sql: str) -> pl.DataFrame:
    with duckdb.connect(database=":memory:") as con:
        con.register("remittances", table)
        return pl.from_arrow(con.execute(sql).to_arrow_table())


def format_money(value: float | int | None) -> str:
    return "0.00" if value is None else f"{float(value):,.2f}"


def table_column_config(columns: list[str]) -> dict[str, object]:
    config: dict[str, object] = {}
    for col in columns:
        if col in MONEY_COLUMNS:
            config[col] = st.column_config.NumberColumn(col, format="%.2f")
        elif col in PERCENT_COLUMNS:
            config[col] = st.column_config.NumberColumn(col, format="%.2f%%")
        elif col in DATE_DISPLAY_COLUMNS:
            config[col] = st.column_config.DateColumn(col, format="YYYY-MM-DD")
        elif col in NUMBER_COLUMNS:
            config[col] = st.column_config.NumberColumn(col)
    return config


def style_balance_rows(row):
    balance = row.get("RMT_BAL", row.get("total_balance", 0))
    try:
        is_balance = abs(float(balance or 0)) > 0.01
    except (TypeError, ValueError):
        is_balance = False
    if is_balance:
        return ["background-color: #fff3cd; color: #5f3900; font-weight: 700"] * len(row)
    return [""] * len(row)


def render_table(data, *, height: int = 480, highlight_balance: bool = True) -> None:
    display = data
    if highlight_balance and hasattr(data, "columns") and not data.empty:
        has_balance = bool({"RMT_BAL", "total_balance"} & set(data.columns))
        if has_balance and data.shape[0] * data.shape[1] <= MAX_STYLED_CELLS:
            display = data.style.apply(style_balance_rows, axis=1)

    st.dataframe(
        display,
        use_container_width=True,
        hide_index=True,
        height=height,
        column_config=table_column_config(list(data.columns)) if hasattr(data, "columns") else None,
    )


def render_download(label: str, data, file_name: str) -> None:
    st.download_button(
        label,
        data.to_csv(index=False).encode("utf-8"),
        file_name=file_name,
        mime="text/csv",
        use_container_width=True,
    )


def style_chart(fig):
    fig.update_layout(margin=dict(l=10, r=10, t=60, b=10), title_x=0.02, legend_title_text="")
    return fig


def safe_columns(df: pl.DataFrame, columns: list[str]) -> list[str]:
    return [col for col in columns if col in df.columns]


def apply_sidebar_filters(df: pl.DataFrame) -> pl.DataFrame:
    filtered = df
    with st.sidebar:
        st.header("Filters")

        if "RMT_DAT" in filtered.columns:
            date_row = filtered.select(
                pl.col("RMT_DAT").min().alias("min_date"),
                pl.col("RMT_DAT").max().alias("max_date"),
            ).row(0, named=True)
            min_date = date_row["min_date"].date() if hasattr(date_row["min_date"], "date") else date_row["min_date"]
            max_date = date_row["max_date"].date() if hasattr(date_row["max_date"], "date") else date_row["max_date"]
            if min_date and max_date:
                selected = st.date_input("Remittance date", value=(min_date, max_date))
                if isinstance(selected, tuple) and len(selected) == 2:
                    start, end = selected
                    filtered = filtered.filter(pl.col("RMT_DAT").dt.date().is_between(start, end))

        if "RMT_CUR" in filtered.columns:
            currencies = filtered.select(pl.col("RMT_CUR").drop_nulls().unique().sort()).to_series().to_list()
            selected_currencies = st.multiselect("Currency", currencies, default=currencies)
            if selected_currencies:
                filtered = filtered.filter(pl.col("RMT_CUR").is_in(selected_currencies))

        search_text = st.text_input("Search current dataset")
        if search_text:
            expressions = [
                pl.col(col).cast(pl.Utf8).str.to_lowercase().str.contains(search_text.lower(), literal=True)
                for col in safe_columns(filtered, TEXT_SEARCH_COLUMNS)
            ]
            if expressions:
                condition = expressions[0]
                for expr in expressions[1:]:
                    condition = condition | expr
                filtered = filtered.filter(condition.fill_null(False))

    return filtered


def search_where_clause(search_text: str, fields: list[str]) -> str:
    if not search_text:
        return ""
    joined_fields = ", ".join(f"cast({field} as varchar)" for field in fields)
    return f"where lower(concat_ws(' ', {joined_fields})) like {sql_like(search_text)}"


# ---------------------------------------------------------------------------
# Report 1: balance remittances
# ---------------------------------------------------------------------------

def show_balance_remittances(table: pa.Table) -> None:
    st.subheader("All balance remittances")

    filter_cols = st.columns([1, 1, 1])
    cnm_search = filter_cols[0].text_input("Search CNM (RMT_CNM)", key="balance_cnm_search")
    bnm_search = filter_cols[1].text_input("Search BNM (RMT_BNM)", key="balance_bnm_search")
    sort_choice = filter_cols[2].selectbox(
        "Sort by",
        [
            "Most remaining balance",
            "Largest absolute balance",
            "Highest remittance amount",
            "Newest remittance date",
            "Oldest remittance date",
        ],
        key="balance_sort",
    )

    where_parts = ["abs(coalesce(RMT_BAL, 0)) > 0.01"]
    if cnm_search:
        where_parts.append(f"lower(cast(RMT_CNM as varchar)) like {sql_like(cnm_search)}")
    if bnm_search:
        where_parts.append(f"lower(cast(RMT_BNM as varchar)) like {sql_like(bnm_search)}")
    balance_where = " and ".join(where_parts)

    sort_sql = {
        "Most remaining balance": "RMT_BAL desc nulls last, RMT_AMT desc nulls last",
        "Largest absolute balance": "abs(RMT_BAL) desc nulls last, RMT_AMT desc nulls last",
        "Highest remittance amount": "RMT_AMT desc nulls last, abs(RMT_BAL) desc nulls last",
        "Newest remittance date": "RMT_DAT desc nulls last, abs(RMT_BAL) desc nulls last",
        "Oldest remittance date": "RMT_DAT asc nulls last, abs(RMT_BAL) desc nulls last",
    }[sort_choice]

    metrics = run_query(
        table,
        f"""
        select
            count(*) as balance_records,
            sum(coalesce(RMT_BAL, 0)) as total_balance,
            max(RMT_BAL) as max_balance,
            sum(coalesce(RMT_AMT, 0)) as balance_remittance_amount
        from remittances
        where {balance_where}
        """,
    ).row(0, named=True)

    cols = st.columns(4)
    cols[0].metric("Balance records", f"{metrics['balance_records'] or 0:,}")
    cols[1].metric("Total RMT_BAL", format_money(metrics["total_balance"]))
    cols[2].metric("Highest RMT_BAL", format_money(metrics["max_balance"]))
    cols[3].metric("Remittance amount", format_money(metrics["balance_remittance_amount"]))

    limit = st.slider("Rows to show", 100, 5000, 1000, 100, key="balance_rows")
    balance_df = run_query(
        table,
        f"""
        select
            RMT_DAT,
            RMT_CNM,
            RMT_BNM,
            RMT_BCN,
            RMT_CUR,
            RMT_AMT,
            RMT_OUT,
            RMT_BAL,
            RMT_TPX,
            RMT_REF,
            IDE_REG_NBR,
            IDE_REG_DAT,
            CMP_CON_COD,
            CMP_CON_NAM,
            DEC_COD,
            DEC_NAM
        from remittances
        where {balance_where}
        order by {sort_sql}
        limit {int(limit)}
        """,
    ).to_pandas()

    left, right = st.columns(2)
    by_currency = run_query(
        table,
        f"""
        select coalesce(RMT_CUR, 'Unknown') as currency,
               count(*) as records,
               sum(coalesce(RMT_BAL, 0)) as total_balance
        from remittances
        where {balance_where}
        group by 1
        order by total_balance desc
        """,
    ).to_pandas()
    if not by_currency.empty:
        fig = px.bar(by_currency, x="currency", y="total_balance", title="Balance by currency")
        style_chart(fig)
        left.plotly_chart(fig, use_container_width=True)

    by_month = run_query(
        table,
        f"""
        select date_trunc('month', RMT_DAT) as month,
               count(*) as records,
               sum(coalesce(RMT_BAL, 0)) as total_balance
        from remittances
        where {balance_where} and RMT_DAT is not null
        group by 1
        order by 1
        """,
    ).to_pandas()
    if not by_month.empty:
        fig = px.line(by_month, x="month", y="total_balance", markers=True, title="Balance trend")
        style_chart(fig)
        right.plotly_chart(fig, use_container_width=True)

    render_table(balance_df)
    render_download("Download displayed balance report", balance_df, "balance_remittances.csv")


# ---------------------------------------------------------------------------
# Report 2: balance remittances grouped by exporter / beneficiary name
# ---------------------------------------------------------------------------

def show_balance_by_exporter(table: pa.Table) -> None:
    st.subheader("Balance remittances grouped by exporter / BNM")

    search = st.text_input("Search exporter (RMT_BNM)", key="bnm_balance_search")
    where_search = ""
    if search:
        where_search = f"and lower(cast(RMT_BNM as varchar)) like {sql_like(search)}"

    grouped = run_query(
        table,
        f"""
        select
            coalesce(RMT_BNM, 'Unknown') as RMT_BNM,
            coalesce(RMT_BCN, 'Unknown') as exporter_country,
            count(*) as records,
            count(distinct RMT_CNM) as remitters,
            count(distinct RMT_REF) as references,
            sum(coalesce(RMT_AMT, 0)) as total_amount,
            sum(coalesce(RMT_OUT, 0)) as total_out,
            sum(coalesce(RMT_BAL, 0)) as total_balance,
            max(RMT_BAL) as max_balance,
            min(RMT_DAT) as first_rmt_date,
            max(RMT_DAT) as last_rmt_date
        from remittances
        where abs(coalesce(RMT_BAL, 0)) > 0.01
        {where_search}
        group by 1, 2
        order by abs(total_balance) desc, total_amount desc
        limit 1000
        """,
    ).to_pandas()

    if grouped.empty:
        st.info("No balance exporters match the current filter.")
        return

    fig = px.bar(
        grouped.head(25),
        x="total_balance",
        y="RMT_BNM",
        color="exporter_country",
        orientation="h",
        title="Top exporters by open balance",
        hover_data={"records": True, "remitters": True, "total_amount": ":,.2f"},
    )
    style_chart(fig)
    fig.update_layout(yaxis_title="", xaxis_title="Total RMT_BAL")
    st.plotly_chart(fig, use_container_width=True)

    render_table(grouped)
    render_download("Download exporter balance summary", grouped, "balance_by_exporter_bnm.csv")


# ---------------------------------------------------------------------------
# Report 3: no CusDec / not-written-off remittances with risk indicators
# ---------------------------------------------------------------------------


def normalize_name_sql(field: str) -> str:
    return f"""
    regexp_replace(
      regexp_replace(
        lower(trim(coalesce(cast({field} as varchar),''))),
        '(\\b(ltd|limited|pvt|private|inc|company|co)\\b)',' ','g'
      ),
      '[^a-z0-9 ]',' ','g'
    )
    """

def no_cusdec_query(category: str, search: str, limit: int) -> str:
    category_filter = "" if category == "All" else f"and category = {sql_string(category)}"
    search_filter = ""
    if search:
        search_filter = f"""
        and lower(concat_ws(' ',
            cast(RMT_CNM as varchar),
            cast(RMT_BNM as varchar),
            cast(RMT_REF as varchar),
            cast(RMT_DES as varchar),
            cast(RMT_BCN as varchar)
        )) like {sql_like(search)}
        """

    return f"""
    with base as (
        select
            *,
            {normalize_name_sql('RMT_BNM')} as normalized_bnm,
            {normalize_name_sql('RMT_CNM')} as normalized_cnm,
            coalesce(RMT_BAL, 0) as open_balance,
            case when RMT_AMT is null or abs(RMT_AMT) = 0 then null
                 else abs(coalesce(RMT_BAL, 0)) / abs(RMT_AMT)
            end as residual_ratio,
            date_diff('day', RMT_DAT, current_date) as days_unwritten
        from remittances
        where {NO_CUSDEC_CONDITION}
    ),
    counted as (
        select
            *,
            count(*) over (partition by RMT_BNM) as bnm_no_cusdec_count,
            count(*) over (partition by RMT_CNM) as cnm_no_cusdec_count,
            count(*) over (partition by RMT_BNM, RMT_CNM) as pair_no_cusdec_count,
            count(*) over (partition by RMT_BNM, RMT_DAT) as same_day_beneficiary_transfers
        from base
    ),
    scored as (
        select
            case
                when normalized_bnm = normalized_cnm and normalized_bnm<>'' then 100
                when normalized_bnm like '%' || normalized_cnm || '%' 
                     or normalized_cnm like '%' || normalized_bnm || '%' then 80
                when levenshtein(normalized_bnm, normalized_cnm) <=2 then 60
                else 30
            end as match_score,

            case
                when normalized_bnm = normalized_cnm and normalized_bnm<>'' then 'Exact Match'
                when normalized_bnm like '%' || normalized_cnm || '%' 
                     or normalized_cnm like '%' || normalized_bnm || '%' then 'Partial Match'
                when levenshtein(normalized_bnm, normalized_cnm) <=2 then 'Spelling Variation'
                else 'Manual Review'
            end as match_type,

            case
                when normalized_bnm = normalized_cnm and normalized_bnm<>'' then 'Exact'
                when (
                     normalized_bnm like '%' || normalized_cnm || '%'
                     or normalized_cnm like '%' || normalized_bnm || '%'
                ) then 'Highly Similar'
                when levenshtein(normalized_bnm, normalized_cnm) <=2 then 'Possible Match'
                else 'Low Similarity'
            end as similarity_category,

            case
                when pair_no_cusdec_count > 1 or (bnm_no_cusdec_count > 1 and cnm_no_cusdec_count > 1) then 'Both'
                when bnm_no_cusdec_count > 1 then 'BNM'
                when cnm_no_cusdec_count > 1 then 'CNM'
                else 'Single'
            end as category,
            (
                25 +
                case when lower(coalesce(RMT_TPX, '')) like '%advanced%' then 15 else 0 end +
                case when abs(open_balance) > 0.01 then 15 else 0 end +
                case when abs(open_balance) >= 10000 then 10 else 0 end +
                case when coalesce(residual_ratio, 0) >= 0.20 then 15 else 0 end +
                case when same_day_beneficiary_transfers >= 3 then 15 else 0 end +
                case when days_unwritten > 90 then 15 else 0 end +
                case when RMT_DES is null
                       or length(trim(cast(RMT_DES as varchar))) <= 3
                       or lower(trim(cast(RMT_DES as varchar))) in ('0', '00', '000', 'goods', 'items', 'material', 'materials', 'sample')
                     then 10 else 0 end +
                case when RMT_AMT >= 100000 then 10 else 0 end
            ) as risk_score,
            concat_ws(' | ',
                'No CusDec / not written off',
                case when lower(coalesce(RMT_TPX, '')) like '%advanced%' then 'Advance remittance without CusDec match' end,
                case when abs(open_balance) > 0.01 then 'Open balance remains' end,
                case when coalesce(residual_ratio, 0) >= 0.20 then 'High residual ratio versus amount' end,
                case when same_day_beneficiary_transfers >= 3 then 'Multiple same-day transfers to same exporter' end,
                case when days_unwritten > 90 then 'Long outstanding remittance' end,
                case when RMT_DES is null
                       or length(trim(cast(RMT_DES as varchar))) <= 3
                       or lower(trim(cast(RMT_DES as varchar))) in ('0', '00', '000', 'goods', 'items', 'material', 'materials', 'sample')
                     then 'Blank or generic goods description' end,
                case when RMT_AMT >= 100000 then 'High value remittance' end
            ) as risk_flags,
            case
                when pair_no_cusdec_count > 1 then 'Repeated exporter-importer pair; review for split remittance, unmatched imports, or valuation suppression.'
                when bnm_no_cusdec_count > 1 and cnm_no_cusdec_count > 1 then 'Repeated exporter and importer pattern; review related parties, split payments, and missing import write-off.'
                when bnm_no_cusdec_count > 1 then 'Repeated exporter with no CusDec; review supplier-level unmatched remittances and possible over-remittance.'
                when cnm_no_cusdec_count > 1 then 'Repeated importer with no CusDec; review importer-level non-write-off behavior and possible informal import settlement.'
                else 'Single no-CusDec remittance; verify documents before treating as irregular.'
            end as possible_review_angle,
            *
        from counted
    )
    select
        category,
        risk_score,
        risk_flags,
        possible_review_angle,
        INSTANCE_ID,
        RMT_DAT,
        days_unwritten,
        RMT_CNM,
        normalized_cnm,
        RMT_BNM,
        normalized_bnm,
        match_score,
        match_type,
        similarity_category,
        RMT_BCN,
        RMT_CUR,
        RMT_AMT,
        RMT_OUT,
        RMT_BAL,
        residual_ratio,
        RMT_TPX,
        RMT_TOD,
        RMT_DES,
        RMT_REF,
        IDE_REG_NBR,
        IDE_REG_DAT,
        CMP_CON_COD,
        CMP_CON_NAM,
        DEC_COD,
        DEC_NAM,
        bnm_no_cusdec_count,
        cnm_no_cusdec_count,
        pair_no_cusdec_count,
        same_day_beneficiary_transfers
    from scored
    where 1 = 1
    {category_filter}
    {search_filter}
    order by risk_score desc, days_unwritten desc nulls last, RMT_AMT desc
    limit {int(limit)}
    """


def show_no_cusdec_report(table: pa.Table) -> None:
    st.subheader("No CusDec / not-written-off remittances")
    st.caption(
        "These are records with no clear customs declaration link. The risk flags are review indicators, "
        "not legal conclusions. They help prioritize possible undervaluation, split-remittance, and non-write-off patterns."
    )

    metrics = run_query(
        table,
        f"""
        select
            count(*) as records,
            count(distinct RMT_BNM) as exporters,
            count(distinct RMT_CNM) as remitters,
            sum(coalesce(RMT_AMT, 0)) as total_amount,
            sum(coalesce(RMT_BAL, 0)) as total_balance
        from remittances
        where {NO_CUSDEC_CONDITION}
        """,
    ).row(0, named=True)

    cols = st.columns(5)
    cols[0].metric("No CusDec records", f"{metrics['records'] or 0:,}")
    cols[1].metric("Exporters (BNM)", f"{metrics['exporters'] or 0:,}")
    cols[2].metric("Remitters (CNM)", f"{metrics['remitters'] or 0:,}")
    cols[3].metric("Amount", format_money(metrics["total_amount"]))
    cols[4].metric("Open balance", format_money(metrics["total_balance"]))

    filter_col, search_col, limit_col = st.columns((0.8, 1.4, 0.8))
    category = filter_col.selectbox("Category", ["All", "Both", "BNM", "CNM", "Single"])
    search = search_col.text_input("Search no-CusDec records", placeholder="Exporter, remitter, reference, country")
    limit = limit_col.slider("Rows", 100, 5000, 1000, 100, key="no_cusdec_rows")

    details = run_query(table, no_cusdec_query(category, search, limit)).to_pandas()
    if details.empty:
        st.info("No no-CusDec records match the current filters.")
        return

    left, right = st.columns(2)
    category_counts = (
        details.groupby("category", as_index=False)
        .agg(records=("INSTANCE_ID", "count"), total_amount=("RMT_AMT", "sum"))
        .sort_values("records", ascending=False)
    )
    fig = px.bar(category_counts, x="category", y="records", color="category", title="No-CusDec category split")
    style_chart(fig)
    left.plotly_chart(fig, use_container_width=True)

    top_bnm = (
        details.groupby("RMT_BNM", as_index=False)
        .agg(records=("INSTANCE_ID", "count"), amount=("RMT_AMT", "sum"), max_risk=("risk_score", "max"))
        .sort_values(["max_risk", "amount"], ascending=[False, False])
        .head(20)
    )
    fig = px.bar(
        top_bnm,
        x="amount",
        y="RMT_BNM",
        color="max_risk",
        orientation="h",
        title="Top no-CusDec exporters",
        hover_data={"records": True, "max_risk": True, "amount": ":,.2f"},
    )
    style_chart(fig)
    right.plotly_chart(fig, use_container_width=True)

    render_table(details, height=560)
    render_download("Download no-CusDec review report", details, "no_cusdec_not_written_off_review.csv")


# ---------------------------------------------------------------------------
# Report 4: optimized consignee-declarant mapping
# ---------------------------------------------------------------------------

# def mapping_where_clause(search: str, declarant: str | None, consignee: str | None) -> str:
#     conditions = []
#     if search:
#         conditions.append(
#             f"""
#             lower(concat_ws(' ',
#                 cast(DEC_COD as varchar),
#                 cast(DEC_NAM as varchar),
#                 cast(CMP_CON_COD as varchar),
#                 cast(CMP_CON_NAM as varchar)
#             )) like {sql_like(search)}
#             """
#         )
#     if declarant:
#         conditions.append(f"{DECLARANT_NODE_SQL} = {sql_string(declarant)}")
#     if consignee:
#         conditions.append(f"{CONSIGNEE_NODE_SQL} = {sql_string(consignee)}")
#     return "" if not conditions else "where " + "\nand ".join(conditions)


# def mapping_dropdown_options(table: pa.Table, node_sql: str, where_clause: str) -> list[str]:
#     result = run_query(
#         table,
#         f"""
#         select {node_sql} as node_name,
#                count(*) as records,
#                sum(coalesce(RMT_AMT, 0)) as total_amount
#         from remittances
#         {where_clause}
#         group by 1
#         order by total_amount desc, records desc
#         limit 250
#         """,
#     ).to_pandas()
#     return [] if result.empty else result["node_name"].dropna().tolist()


# def bloom_text(value: object, max_chars: int = 22) -> str:
#     text = "Unknown" if value is None else str(value).replace("\n", " ").strip() or "Unknown"
#     if len(text) > max_chars:
#         text = f"{text[: max_chars - 3]}..."
#     parts = text.split()
#     if len(parts) > 1:
#         midpoint = max(1, len(parts) // 2)
#         return " ".join(parts[:midpoint]) + "<br>" + " ".join(parts[midpoint:])
#     return text


# def build_mapping_network(mapping, center_mode: str, link_limit: int):
#     center_col = "consignee_node" if center_mode == "Consignee" else "declarant_node"
#     center_name = "consignee_name" if center_mode == "Consignee" else "dec_name"
#     outer_col = "declarant_node" if center_mode == "Consignee" else "consignee_node"
#     outer_name = "dec_name" if center_mode == "Consignee" else "consignee_name"

#     focus_source = mapping[mapping[center_col] != "Unknown"]
#     if focus_source.empty:
#         focus_source = mapping

#     focus = (
#         focus_source.groupby([center_col, center_name], as_index=False)
#         .agg(transfers=("transfers", "sum"), total_amount=("total_amount", "sum"), outer_count=(outer_col, "nunique"))
#         .sort_values(["outer_count", "transfers", "total_amount"], ascending=[False, False, False])
#         .iloc[0]
#     )

#     center_node = focus[center_col]
#     links = mapping[mapping[center_col] == center_node].sort_values(
#         ["total_amount", "transfers"], ascending=[False, False]
#     ).head(link_limit)

#     radius = 2.5
#     count = len(links)
#     angle_step = 2 * math.pi / max(count, 1)

#     x_values = [0]
#     y_values = [0]
#     labels = [bloom_text(focus[center_name])]
#     hover = [f"{center_node}<br>Transfers: {int(focus['transfers'])}<br>Amount: {focus['total_amount']:,.2f}"]
#     sizes = [78]
#     colors = ["#5aa38e"]

#     fig = go.Figure()
#     for idx, row in enumerate(links.itertuples(index=False)):
#         angle = math.pi / 2 + idx * angle_step
#         x = radius * math.cos(angle)
#         y = radius * math.sin(angle)
#         x_values.append(x)
#         y_values.append(y)
#         labels.append(bloom_text(getattr(row, outer_name)))
#         hover.append(
#             f"{getattr(row, outer_col)}<br>"
#             f"Transfers: {int(row.transfers)}<br>"
#             f"References: {int(row.references)}<br>"
#             f"Amount: {row.total_amount:,.2f}<br>"
#             f"Balance: {row.total_balance:,.2f}"
#         )
#         sizes.append(max(46, min(84, 42 + math.sqrt(max(float(row.total_amount), 0)) / 90)))
#         colors.append("#d86491")

#         fig.add_annotation(
#             x=x * 0.82,
#             y=y * 0.82,
#             ax=0.25 * math.cos(angle),
#             ay=0.25 * math.sin(angle),
#             xref="x",
#             yref="y",
#             axref="x",
#             ayref="y",
#             showarrow=True,
#             arrowhead=3,
#             arrowsize=1.1,
#             arrowwidth=2,
#             arrowcolor="#aab2bd",
#         )
#         fig.add_annotation(
#             x=x * 0.5,
#             y=y * 0.5,
#             text="DECLARANT_OF",
#             showarrow=False,
#             font=dict(color="#d1d5db", size=10),
#             textangle=math.degrees(angle),
#         )

#     fig.add_trace(
#         go.Scatter(
#             x=x_values,
#             y=y_values,
#             mode="markers+text",
#             text=labels,
#             hovertext=hover,
#             hoverinfo="text",
#             textposition="middle center",
#             textfont=dict(color="#17212b", size=11),
#             marker=dict(size=sizes, color=colors, line=dict(color="rgba(255,255,255,0.12)", width=1)),
#         )
#     )
#     fig.update_layout(
#         title_text=f"{center_mode}-centered consignee-declarant mapping",
#         height=680,
#         paper_bgcolor="#000000",
#         plot_bgcolor="#000000",
#         showlegend=False,
#         margin=dict(l=10, r=10, t=60, b=10),
#         xaxis=dict(visible=False, range=[-3.25, 3.25], fixedrange=True),
#         yaxis=dict(visible=False, range=[-3, 3], fixedrange=True, scaleanchor="x", scaleratio=1),
#     )
#     return fig, center_node


# def show_consignee_declarant_mapping(table: pa.Table) -> None:
#     st.subheader("Consignee-declarant mapping")

#     search_col, mode_col, link_col, row_col = st.columns((1.6, 0.8, 0.7, 0.8))
#     search = search_col.text_input("Search mapping", placeholder="Declarant, consignee, code, or name")
#     center_mode = mode_col.radio("Center", ["Consignee", "Declarant"], horizontal=True)
#     link_limit = link_col.slider("Node links", 5, 50, 20, 5)
#     row_limit = row_col.slider("Rows", 100, 5000, 1000, 100)

#     option_where = mapping_where_clause(search, None, None)
#     declarant_options = mapping_dropdown_options(table, DECLARANT_NODE_SQL, option_where)
#     consignee_options = mapping_dropdown_options(table, CONSIGNEE_NODE_SQL, option_where)

#     left, right = st.columns(2)
#     selected_declarant_label = left.selectbox("Declarant dropdown", ["All declarants", *declarant_options])
#     selected_consignee_label = right.selectbox("Consignee dropdown", ["All consignees", *consignee_options])
#     selected_declarant = None if selected_declarant_label == "All declarants" else selected_declarant_label
#     selected_consignee = None if selected_consignee_label == "All consignees" else selected_consignee_label

#     where_clause = mapping_where_clause(search, selected_declarant, selected_consignee)
#     mapping = run_query(
#         table,
#         f"""
#         select
#             coalesce(DEC_COD, 'Unknown') as dec_cod,
#             coalesce(DEC_NAM, 'Unknown') as dec_name,
#             coalesce(CMP_CON_COD, 'Unknown') as consignee_code,
#             coalesce(CMP_CON_NAM, 'Unknown') as consignee_name,
#             {DECLARANT_NODE_SQL} as declarant_node,
#             {CONSIGNEE_NODE_SQL} as consignee_node,
#             count(*) as transfers,
#             count(distinct RMT_REF) as references,
#             sum(coalesce(RMT_AMT, 0)) as total_amount,
#             sum(coalesce(RMT_BAL, 0)) as total_balance,
#             min(RMT_DAT) as first_rmt_date,
#             max(RMT_DAT) as last_rmt_date
#         from remittances
#         {where_clause}
#         group by 1, 2, 3, 4, 5, 6
#         order by total_amount desc, transfers desc
#         limit {int(row_limit)}
#         """,
#     ).to_pandas()

#     if mapping.empty:
#         st.info("No mapping records match the current filters.")
#         return

#     fig, center_node = build_mapping_network(mapping, center_mode, link_limit)
#     st.plotly_chart(fig, use_container_width=True)
#     st.caption(f"Centered on: {center_node}")

#     fig = px.bar(
#         mapping.head(30),
#         x="total_amount",
#         y="consignee_node" if center_mode == "Consignee" else "declarant_node",
#         color="declarant_node" if center_mode == "Consignee" else "consignee_node",
#         orientation="h",
#         title="Top consignee-declarant links",
#         hover_data={"transfers": True, "references": True, "total_amount": ":,.2f"},
#     )
#     style_chart(fig)
#     st.plotly_chart(fig, use_container_width=True)

#     display = mapping.drop(columns=["declarant_node", "consignee_node"], errors="ignore")
#     render_table(display)
#     render_download("Download consignee-declarant mapping", display, "consignee_declarant_mapping.csv")


def show_data_quality(table: pa.Table) -> None:
    st.subheader("Data quality snapshot")
    quality = run_query(
        table,
        f"""
        select
            count(*) as records,
            count(*) filter (where abs(coalesce(RMT_BAL, 0)) > 0.01) as balance_records,
            count(*) filter (where {NO_CUSDEC_CONDITION}) as no_cusdec_records,
            count(distinct RMT_BNM) as exporters,
            count(distinct RMT_CNM) as remitters,
            count(distinct DEC_COD) as declarants,
            count(distinct CMP_CON_COD) as consignees,
            sum(coalesce(RMT_AMT, 0)) as total_amount,
            sum(coalesce(RMT_BAL, 0)) as total_balance
        from remittances
        """,
    ).to_pandas()
    render_table(quality, height=120)


def mapping_where_clause(search: str, declarant: str | None, consignee: str | None) -> str:
    conditions = []

    if search:
        conditions.append(
            f"""
            lower(concat_ws(' ',
                cast(DEC_COD as varchar),
                cast(DEC_NAM as varchar),
                cast(CMP_CON_COD as varchar),
                cast(CMP_CON_NAM as varchar)
            )) like {sql_like(search)}
            """
        )

    if declarant:
        conditions.append(
            f"{DECLARANT_NODE_SQL}={sql_string(declarant)}"
        )

    if consignee:
        conditions.append(
            f"{CONSIGNEE_NODE_SQL}={sql_string(consignee)}"
        )

    return "" if not conditions else "where " + "\nand ".join(conditions)


def mapping_dropdown_options(
    table: pa.Table,
    node_sql: str,
    where_clause: str
):

    result = run_query(
        table,
        f"""
        select
            {node_sql} as node_name,
            count(*) records,
            sum(coalesce(RMT_AMT,0)) total_amount
        from remittances
        {where_clause}
        group by 1
        order by total_amount desc,records desc
        limit 250
        """
    ).to_pandas()

    if result.empty:
        return []

    return result["node_name"].dropna().tolist()


def bloom_text(
    value: object,
    max_chars: int = 22
):

    text = (
        "Unknown"
        if value is None
        else str(value).replace("\n"," ").strip()
    )

    if not text:
        text="Unknown"

    if len(text)>max_chars:
        text=text[:max_chars-3]+"..."

    words=text.split()

    if len(words)>1:
        midpoint=max(1,len(words)//2)

        return (
            " ".join(words[:midpoint])
            +"<br>"
            +" ".join(words[midpoint:])
        )

    return text


def build_mapping_network(
    mapping,
    link_limit: int
):

    links=(
        mapping
        .sort_values(
            ["total_amount","transfers"],
            ascending=[False,False]
        )
        .head(link_limit)
    )

    fig=go.Figure()

    node_positions={}
    nodes=set()

    for row in links.itertuples(index=False):

        nodes.add(row.consignee_node)
        nodes.add(row.declarant_node)

    nodes=list(nodes)

    radius=3

    angle_step=(
        2*math.pi/max(len(nodes),1)
    )

    for i,node in enumerate(nodes):

        angle=i*angle_step

        x=radius*math.cos(angle)
        y=radius*math.sin(angle)

        node_positions[node]=(x,y)

    # Draw relationship lines

    for row in links.itertuples(index=False):

        x0,y0=node_positions[row.consignee_node]
        x1,y1=node_positions[row.declarant_node]

        fig.add_trace(
            go.Scatter(
                x=[x0,x1],
                y=[y0,y1],
                mode="lines",
                line=dict(
                    width=max(
                        1,
                        min(
                            8,
                            math.sqrt(
                                max(
                                    float(
                                        row.total_amount
                                    ),
                                    1
                                )
                            )/100
                        )
                    )
                ),
                hoverinfo="text",
                text=
                f"{row.consignee_node}"
                f"<br>→ {row.declarant_node}"
                f"<br>Transfers:{row.transfers}"
                # f"<br>References:{row.references}"
                f"<br>References:{row.ref_count}"
                f"<br>Amount:{row.total_amount:,.2f}"
                f"<br>Balance:{row.total_balance:,.2f}"
            )
        )

    xvals=[]
    yvals=[]
    labels=[]
    hover=[]
    sizes=[]

    for node in nodes:

        x,y=node_positions[node]

        subset=links[
            (
                links["consignee_node"]==node
            )
            |
            (
                links["declarant_node"]==node
            )
        ]

        amount=subset[
            "total_amount"
        ].sum()

        xvals.append(x)
        yvals.append(y)

        labels.append(
            bloom_text(node)
        )

        hover.append(
            f"{node}"
            f"<br>Total amount:{amount:,.2f}"
            f"<br>Connections:{len(subset)}"
        )

        sizes.append(
            max(
                40,
                min(
                    90,
                    40+
                    math.sqrt(
                        max(amount,1)
                    )/80
                )
            )
        )

    fig.add_trace(
        go.Scatter(
            x=xvals,
            y=yvals,
            mode="markers+text",
            text=labels,
            hovertext=hover,
            hoverinfo="text",
            textposition="middle center",
            textfont=dict(
                color="#17212b",
                size=11
            ),
            marker=dict(
                size=sizes,
                color="#5aa38e",
                line=dict(
                    color="rgba(255,255,255,0.12)",
                    width=1
                )
            )
        )
    )

    fig.update_layout(
        title="Consignee ↔ Declarant Node Mapping",
        height=700,
        paper_bgcolor="#000000",
        plot_bgcolor="#000000",
        showlegend=False,
        margin=dict(
            l=10,
            r=10,
            t=60,
            b=10
        ),
        xaxis=dict(
            visible=False
        ),
        yaxis=dict(
            visible=False,
            scaleanchor="x",
            scaleratio=1
        )
    )

    return fig


# def show_consignee_declarant_mapping(
#     table: pa.Table
# )->None:

#     st.subheader(
#         "Consignee-declarant mapping"
#     )

#     search_col,link_col,row_col=st.columns(
#         (1.8,0.7,0.8)
#     )

#     search=search_col.text_input(
#         "Search mapping",
#         placeholder="Declarant, consignee, code or name"
#     )

#     link_limit=link_col.slider(
#         "Node links",
#         5,
#         50,
#         20,
#         5
#     )

#     row_limit=row_col.slider(
#         "Rows",
#         100,
#         5000,
#         1000,
#         100
#     )

#     option_where=(
#         mapping_where_clause(
#             search,
#             None,
#             None
#         )
#     )

#     declarant_options=(
#         mapping_dropdown_options(
#             table,
#             DECLARANT_NODE_SQL,
#             option_where
#         )
#     )

#     consignee_options=(
#         mapping_dropdown_options(
#             table,
#             CONSIGNEE_NODE_SQL,
#             option_where
#         )
#     )

#     left,right=st.columns(2)

#     selected_declarant_label=left.selectbox(
#         "Declarant dropdown",
#         ["All declarants",*declarant_options]
#     )

#     selected_consignee_label=right.selectbox(
#         "Consignee dropdown",
#         ["All consignees",*consignee_options]
#     )

#     selected_declarant=(
#         None
#         if selected_declarant_label=="All declarants"
#         else selected_declarant_label
#     )

#     selected_consignee=(
#         None
#         if selected_consignee_label=="All consignees"
#         else selected_consignee_label
#     )

#     where_clause=(
#         mapping_where_clause(
#             search,
#             selected_declarant,
#             selected_consignee
#         )
#     )

#     mapping=run_query(
#         table,
#         f"""
#         select

#         {DECLARANT_NODE_SQL}
#         as declarant_node,

#         {CONSIGNEE_NODE_SQL}
#         as consignee_node,

#         count(*) transfers,

#         count(distinct RMT_REF)
#         references,

#         sum(coalesce(RMT_AMT,0))
#         total_amount,

#         sum(coalesce(RMT_BAL,0))
#         total_balance,

#         min(RMT_DAT)
#         first_rmt_date,

#         max(RMT_DAT)
#         last_rmt_date

#         from remittances

#         {where_clause}

#         group by 1,2

#         order by
#         total_amount desc,
#         transfers desc

#         limit {int(row_limit)}
#         """
#     ).to_pandas()

#     if mapping.empty:

#         st.info(
#             "No mapping records match current filters."
#         )
#         return

#     fig=build_mapping_network(
#         mapping,
#         link_limit
#     )

#     st.plotly_chart(
#         fig,
#         use_container_width=True
#     )

#     fig=px.bar(
#         mapping.head(30),
#         x="total_amount",
#         y="consignee_node",
#         color="declarant_node",
#         orientation="h",
#         title="Top node relationships",
#         hover_data={
#             "transfers":True,
#             "references":True,
#             "total_amount":":,.2f"
#         }
#     )

#     style_chart(fig)

#     st.plotly_chart(
#         fig,
#         use_container_width=True
#     )

#     render_table(
#         mapping
#     )

#     render_download(
#         "Download consignee-declarant mapping",
#         mapping,
#         "consignee_declarant_mapping.csv"
#     )

#working GGCODE - 2026.05.21 
# def show_consignee_declarant_mapping(
#     table: pa.Table
# ) -> None:

#     st.subheader(
#         "Consignee-declarant mapping"
#     )

#     search_col, link_col, row_col = st.columns(
#         (1.8, 0.7, 0.8)
#     )

#     search = search_col.text_input(
#         "Search mapping",
#         placeholder="Declarant, consignee, code or name"
#     )

#     link_limit = link_col.slider(
#         "Node links",
#         5,
#         50,
#         20,
#         5
#     )

#     row_limit = row_col.slider(
#         "Rows",
#         100,
#         5000,
#         1000,
#         100
#     )

#     option_where = mapping_where_clause(
#         search,
#         None,
#         None
#     )

#     declarant_options = mapping_dropdown_options(
#         table,
#         DECLARANT_NODE_SQL,
#         option_where
#     )

#     consignee_options = mapping_dropdown_options(
#         table,
#         CONSIGNEE_NODE_SQL,
#         option_where
#     )

#     left, right = st.columns(2)

#     selected_declarant_label = left.selectbox(
#         "Declarant dropdown",
#         ["All declarants", *declarant_options]
#     )

#     selected_consignee_label = right.selectbox(
#         "Consignee dropdown",
#         ["All consignees", *consignee_options]
#     )

#     selected_declarant = (
#         None
#         if selected_declarant_label == "All declarants"
#         else selected_declarant_label
#     )

#     selected_consignee = (
#         None
#         if selected_consignee_label == "All consignees"
#         else selected_consignee_label
#     )

#     where_clause = mapping_where_clause(
#         search,
#         selected_declarant,
#         selected_consignee
#     )

#     mapping = run_query(
#         table,
#         f"""
#         select

#         {DECLARANT_NODE_SQL} as declarant_node,

#         {CONSIGNEE_NODE_SQL} as consignee_node,

#         count(*) as transfers,

#         count(distinct RMT_REF) as ref_count,

#         sum(coalesce(RMT_AMT,0)) as total_amount,

#         sum(coalesce(RMT_BAL,0)) as total_balance,

#         min(RMT_DAT) as first_rmt_date,

#         max(RMT_DAT) as last_rmt_date

#         from remittances

#         {where_clause}

#         group by 1,2

#         order by
#         total_amount desc,
#         transfers desc

#         limit {int(row_limit)}
#         """
#     ).to_pandas()

#     if mapping.empty:
#         st.info(
#             "No mapping records match current filters."
#         )
#         return

#     fig = build_mapping_network(
#         mapping,
#         link_limit
#     )

#     st.plotly_chart(
#         fig,
#         use_container_width=True
#     )

#     fig = px.bar(
#         mapping.head(30),
#         x="total_amount",
#         y="consignee_node",
#         color="declarant_node",
#         orientation="h",
#         title="Top node relationships",
#         hover_data={
#             "transfers": True,
#             "ref_count": True,
#             "total_amount": ":,.2f"
#         }
#     )

#     style_chart(fig)

#     st.plotly_chart(
#         fig,
#         use_container_width=True
#     )

#     render_table(
#         mapping
#     )

#     render_download(
#         "Download consignee-declarant mapping",
#         mapping,
#         "consignee_declarant_mapping.csv"
#     )

def show_consignee_declarant_mapping(
    table: pa.Table
) -> None:

    st.subheader("Consignee-declarant mapping")

    search_col, link_col, row_col = st.columns(
        (1.8,0.7,0.8)
    )

    search = search_col.text_input(
        "Search mapping",
        placeholder="Declarant, consignee, code or name"
    )

    link_limit = link_col.slider(
        "Node links",
        5,
        50,
        20,
        5
    )

    row_limit = row_col.slider(
        "Rows",
        100,
        5000,
        1000,
        100
    )

    option_where = mapping_where_clause(
        search,
        None,
        None
    )

    declarant_options = mapping_dropdown_options(
        table,
        DECLARANT_NODE_SQL,
        option_where
    )

    consignee_options = mapping_dropdown_options(
        table,
        CONSIGNEE_NODE_SQL,
        option_where
    )

    left,right = st.columns(2)

    selected_declarant_label = left.selectbox(
        "Declarant dropdown",
        ["All declarants",*declarant_options]
    )

    selected_consignee_label = right.selectbox(
        "Consignee dropdown",
        ["All consignees",*consignee_options]
    )

    selected_declarant = (
        None if selected_declarant_label=="All declarants"
        else selected_declarant_label
    )

    selected_consignee = (
        None if selected_consignee_label=="All consignees"
        else selected_consignee_label
    )

    where_clause = mapping_where_clause(
        search,
        selected_declarant,
        selected_consignee
    )

    mapping = run_query(
        table,
        f"""
        select

        {DECLARANT_NODE_SQL} as declarant_node,

        {CONSIGNEE_NODE_SQL} as consignee_node,

        count(*) as transfers,

        count(distinct RMT_REF) as ref_count,

        sum(coalesce(RMT_AMT,0)) as total_amount,

        sum(coalesce(RMT_BAL,0)) as total_balance,

        min(RMT_DAT) as first_rmt_date,

        max(RMT_DAT) as last_rmt_date

        from remittances

        {where_clause}

        group by 1,2

        order by
        total_amount desc,
        transfers desc

        limit {int(row_limit)}
        """
    ).to_pandas()

    if mapping.empty:
        st.info(
            "No mapping records match current filters."
        )
        return

    fig = build_mapping_network(
        mapping,
        link_limit
    )

    st.plotly_chart(
        fig,
        use_container_width=True
    )

    fig = px.bar(
        mapping.head(30),
        x="total_amount",
        y="consignee_node",
        color="declarant_node",
        orientation="h",
        title="Top node relationships",
        hover_data={
            "transfers":True,
            "ref_count":True,
            "total_amount":":,.2f"
        }
    )

    style_chart(fig)

    st.plotly_chart(
        fig,
        use_container_width=True
    )

    render_table(
        mapping
    )


    # ============================================
    # Balance remittance summary by beneficiary
    # ============================================

    st.subheader(
        "Balance remittances grouped by beneficiary"
    )

    beneficiary_balance = run_query(
        table,
        """
        select

        coalesce(
            RMT_BNM,
            'Unknown'
        ) as beneficiary,

        coalesce(
            RMT_BCN,
            'Unknown'
        ) as beneficiary_country,

        count(*) as records,

        count(distinct RMT_CNM)
        as remitters,

        count(distinct RMT_REF)
        as ref_count,

        sum(coalesce(RMT_AMT,0))
        as total_amount,

        sum(coalesce(RMT_OUT,0))
        as total_out,

        sum(coalesce(RMT_BAL,0))
        as total_balance,

        max(RMT_BAL)
        as max_balance,

        min(RMT_DAT)
        as first_rmt_date,

        max(RMT_DAT)
        as last_rmt_date

        from remittances

        where abs(
            coalesce(RMT_BAL,0)
        )>0.01

        group by 1,2

        order by
        total_balance desc,
        total_amount desc
        """
    ).to_pandas()

    if beneficiary_balance.empty:

        st.info(
            "No beneficiary balance records found."
        )

    else:

        render_table(
            beneficiary_balance,
            height=420
        )

        render_download(
            "Download beneficiary balance summary",
            beneficiary_balance,
            "beneficiary_balance_summary.csv"
        )


    render_download(
        "Download consignee-declarant mapping",
        mapping,
        "consignee_declarant_mapping.csv"
    )
    



BNM_NODE_SQL = """
concat(
coalesce(cast(RMT_BNM as varchar),'Unknown'),
' | ',
left(
regexp_replace(
lower(trim(coalesce(cast(RMT_BNM as varchar),''))),
'(\\b(ltd|limited|pvt|private|inc|company|co)\\b)',' ','g'
),
58
)
)
"""

CNM_NODE_SQL = """
concat(
coalesce(cast(RMT_CNM as varchar),'Unknown'),
' | ',
left(replace(coalesce(cast(RMT_CNM as varchar),'Unknown'),chr(10),' '),58)
)
"""

def show_bnm_cnm_mapping(table: pa.Table)->None:

    st.subheader("BNM-CNM mapping (All Remittances)")

    search_col,link_col,row_col=st.columns((1.8,0.7,0.8))

    search=search_col.text_input(
        "Search BNM/CNM",
        placeholder="Sender or beneficiary"
    )

    link_limit=link_col.slider(
        "Node links",5,50,20,5,
        key="bnmcnm_links"
    )

    row_limit=row_col.slider(
        "Rows",100,5000,1000,100,
        key="bnmcnm_rows"
    )

    where_clause=""

    if search:
        where_clause=f"""
        where lower(concat_ws(' ',
        cast(RMT_CNM as varchar),
        cast(RMT_BNM as varchar)))
        like {sql_like(search)}
        """

    mapping=run_query(
    table,
    f"""
    select

    {CNM_NODE_SQL} as cnm_node,

    {BNM_NODE_SQL} as bnm_node,

    count(*) transfers,

    count(distinct RMT_REF) ref_count,

    count(distinct RMT_CUR) currency_count,

    sum(coalesce(RMT_AMT,0)) total_amount,

    sum(coalesce(RMT_OUT,0)) total_out,

    sum(coalesce(RMT_BAL,0)) total_balance,

    case
    when nullif(sum(coalesce(RMT_BAL,0)),0) is null then null
    else (sum(coalesce(RMT_BAL,0))/nullif(sum(coalesce(RMT_BAL,0)),0))*100
    end as "balance Ratio",

    case
    when nullif(sum(coalesce(RMT_AMT,0)),0) is null then null
    else (sum(coalesce(RMT_BAL,0))/nullif(sum(coalesce(RMT_AMT,0)),0))*100
    end as balance_to_amount_ratio,

    count(*) filter (
    where abs(coalesce(RMT_BAL,0))>0.01
    ) open_balance_transfers,

    case
    when count(*)=0 then null
    else (
    count(*) filter (
    where abs(coalesce(RMT_BAL,0))>0.01
    )::double/count(*)
    )*100
    end as open_balance_rate,

    avg(coalesce(RMT_AMT,0)) avg_amount,

    avg(coalesce(RMT_BAL,0)) avg_balance,

    max(RMT_BAL) max_balance,

    min(RMT_DAT) first_rmt_date,

    max(RMT_DAT) last_rmt_date,

    date_diff('day',min(RMT_DAT),max(RMT_DAT)) active_days

    from remittances

    {where_clause}

    group by 1,2

    order by
    total_amount desc,
    transfers desc

    limit {int(row_limit)}
    """
    ).to_pandas()

    if mapping.empty:
        st.info("No BNM-CNM mapping records found.")
        return

    network=mapping.rename(
    columns={
    "cnm_node":"consignee_node",
    "bnm_node":"declarant_node"
    })

    fig=build_mapping_network(
    network,
    link_limit
    )

    st.plotly_chart(
    fig,
    use_container_width=True
    )

    fig=px.bar(
    mapping.head(30),
    x="total_amount",
    y="bnm_node",
    color="cnm_node",
    orientation="h",
    title="Top BNM-CNM relationships",
    hover_data={
    "transfers":True,
    "ref_count":True
    }
    )

    style_chart(fig)

    st.plotly_chart(fig,use_container_width=True)

    render_table(mapping)

    render_download(
    "Download BNM-CNM mapping",
    mapping,
    "bnm_cnm_mapping.csv"
    )


    # ==========================================================
    # BNM Remittance Data Sheet
    # ==========================================================

    st.markdown("---")
    st.subheader("Comprehensive BNM Remittance Data Sheet")

    c1,c2,c3,c4,c5 = st.columns(5)

    bnm_class_filter = c1.selectbox(
        "BNM Classification",
        ["All"] + sorted(
            run_query(
                table,
                """
                select distinct cast(CMP_CON_COD as varchar)
                from remittances
                where CMP_CON_COD is not null
                """
            ).to_pandas().iloc[:,0].dropna().astype(str).tolist()
        ),
        key="bnm_sheet_class"
    )

    currency_filter = c2.selectbox(
        "Currency",
        ["All"] + sorted(
            run_query(
                table,
                """
                select distinct cast(RMT_CUR as varchar)
                from remittances
                where RMT_CUR is not null
                """
            ).to_pandas().iloc[:,0].dropna().astype(str).tolist()
        ),
        key="bnm_sheet_currency"
    )

    sender_search = c3.text_input(
        "Sender (CNM)",
        key="bnm_sheet_sender"
    )

    recipient_search = c4.text_input(
        "Recipient (BNM)",
        key="bnm_sheet_recipient"
    )

    date_filter = c5.date_input(
        "From Date",
        value=None,
        key="bnm_sheet_date"
    )

    sheet_query = """
    select
    RMT_DAT transaction_date,
    RMT_CNM sender,
    RMT_BNM recipient,
    RMT_AMT amount,
    RMT_CUR currency,
    CMP_CON_COD bnm_classification,
    RMT_REF remittance_reference,
    RMT_BAL balance,
    IDE_REG_NBR customs_reference,
    RMT_BNK bank,
    RMT_BCN country
    from remittances
    where 1=1
    """

    if bnm_class_filter!="All":
        sheet_query += f" and cast(CMP_CON_COD as varchar)='{bnm_class_filter}'"

    if currency_filter!="All":
        sheet_query += f" and cast(RMT_CUR as varchar)='{currency_filter}'"

    if sender_search:
        sheet_query += f" and lower(cast(RMT_CNM as varchar)) like '%{sender_search.lower()}%'"

    if recipient_search:
        sheet_query += f" and lower(cast(RMT_BNM as varchar)) like '%{recipient_search.lower()}%'"

    if date_filter:
        sheet_query += f" and cast(RMT_DAT as date)>='{date_filter}'"

    sheet_query += " order by RMT_DAT desc limit 10000"

    sheet=run_query(table,sheet_query).to_pandas()

    m1,m2,m3,m4=st.columns(4)

    m1.metric("Transactions",f"{len(sheet):,}")
    m2.metric("Total Amount",f"{sheet['amount'].fillna(0).sum():,.2f}")
    m3.metric("Unique Senders",f"{sheet['sender'].nunique():,}")
    m4.metric("Unique Recipients",f"{sheet['recipient'].nunique():,}")

    st.dataframe(
        sheet,
        use_container_width=True,
        height=550
    )

    render_download(
        "Download BNM Remittance Sheet",
        sheet,
        "bnm_remittance_sheet.csv"
    )


def show_consignee_declarant_remittance_amount(table: pa.Table)->None:
    st.subheader("Consignee-declarant remittance amount by customs key")

    search_col, row_col = st.columns((2, 0.7))

    search = search_col.text_input(
        "Search consignee / declarant / customs key",
        placeholder="Consignee, declarant, year, or office code",
        key="consignee_decl_amount_search",
    )

    row_limit = row_col.slider(
        "Rows",
        100,
        5000,
        1000,
        100,
        key="consignee_decl_amount_rows",
    )

    where_clause = ""
    if search:
        where_clause = f"""
        where lower(concat_ws(' ',
            cast(CMP_CON_COD as varchar),
            cast(CMP_CON_NAM as varchar),
            cast(DEC_COD as varchar),
            cast(DEC_NAM as varchar),
            cast(DEC_REF_YER as varchar),
            cast(IDE_CUO_COD as varchar)
        )) like {sql_like(search)}
        """

    summary = run_query(
        table,
        f"""
        select
            coalesce(cast(CMP_CON_COD as varchar),'Unknown') as consignee_code,
            coalesce(cast(CMP_CON_NAM as varchar),'Unknown') as consignee_name,
            coalesce(cast(DEC_COD as varchar),'Unknown') as declarant_code,
            coalesce(cast(DEC_NAM as varchar),'Unknown') as declarant_name,
            coalesce(cast(DEC_REF_YER as varchar),'Unknown') as DEC_REF_YER,
            coalesce(cast(IDE_CUO_COD as varchar),'Unknown') as IDE_CUO_COD,
            concat(
                coalesce(cast(DEC_REF_YER as varchar),'Unknown'),
                '-',
                coalesce(cast(IDE_CUO_COD as varchar),'Unknown')
            ) as customs_composite_key,
            count(*) as remittance_count,
            count(distinct RMT_REF) as ref_count,
            count(distinct RMT_CNM) as remitter_count,
            count(distinct RMT_BNM) as beneficiary_count,
            count(distinct RMT_CUR) as currency_count,
            sum(coalesce(RMT_AMT,0)) as total_sent_amount,
            avg(coalesce(RMT_AMT,0)) as avg_sent_amount,
            sum(coalesce(RMT_BAL,0)) as total_balance,
            sum(coalesce(RMT_OUT,0)) as total_out,
            min(RMT_DAT) as first_rmt_date,
            max(RMT_DAT) as last_rmt_date,
            date_diff('day',min(RMT_DAT),max(RMT_DAT)) as active_days
        from remittances
        {where_clause}
        group by 1,2,3,4,5,6,7
        order by total_sent_amount desc, remittance_count desc
        limit {int(row_limit)}
        """,
    ).to_pandas()

    if summary.empty:
        st.info("No consignee-declarant remittance amount records found.")
        return

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Groups", f"{len(summary):,}")
    m2.metric("Sent amount", format_money(summary["total_sent_amount"].fillna(0).sum()))
    m3.metric("Open balance", format_money(summary["total_balance"].fillna(0).sum()))
    m4.metric("Remittances", f"{summary['remittance_count'].fillna(0).sum():,.0f}")

    top_chart = summary.head(30)
    fig = px.bar(
        top_chart,
        x="total_sent_amount",
        y="consignee_name",
        color="declarant_name",
        orientation="h",
        title="Top consignee-declarant customs keys by remittance sent amount",
        hover_data={
            "customs_composite_key": True,
            "remittance_count": True,
            "ref_count": True,
            "total_balance": ":,.2f",
        },
    )
    style_chart(fig)
    st.plotly_chart(fig, use_container_width=True)

    render_table(summary)
    render_download(
        "Download consignee-declarant remittance amount summary",
        summary,
        "consignee_declarant_remittance_amount.csv",
    )


def show_bnm_vs_imports(table: pa.Table)->None:
    st.subheader("BNM vs imports")

    c1, c2, c3, c4 = st.columns((1.6, 0.9, 0.8, 0.7))
    search = c1.text_input(
        "Search BNM / import key",
        placeholder="BNM, year, office, CusDec, consignee, declarant, or sender",
        key="bnm_vs_imports_search",
    )
    import_status = c2.selectbox(
        "Import status",
        ["All", "With imports only", "No imports only"],
        key="bnm_vs_imports_status",
    )
    min_total_amount = c3.number_input(
        "Min sent amount",
        min_value=0.0,
        value=0.0,
        step=10000.0,
        key="bnm_vs_imports_min_amount",
    )
    row_limit = c4.slider(
        "Rows",
        100,
        5000,
        1000,
        100,
        key="bnm_vs_imports_rows",
    )

    search_fields = [
        "RMT_BNM",
        "RMT_BCN",
        "RMT_CNM",
        "RMT_REF",
        "DEC_REF_YER",
        "IDE_CUO_COD",
        "IDE_REG_NBR",
        "CMP_CON_COD",
        "CMP_CON_NAM",
        "DEC_COD",
        "DEC_NAM",
    ]
    where_conditions = ["1=1"]
    if search:
        joined_fields = ", ".join(f"cast({field} as varchar)" for field in search_fields)
        where_conditions.append(f"lower(concat_ws(' ', {joined_fields})) like {sql_like(search)}")
    if import_status == "With imports only":
        where_conditions.append(f"not {NO_CUSDEC_CONDITION}")
    elif import_status == "No imports only":
        where_conditions.append(NO_CUSDEC_CONDITION)
    where_conditions.append(f"coalesce(cast(IDE_CUO_COD as varchar),'Unknown') <> 'Unknown'")

    where_clause = "where " + "\n        and ".join(f"({condition})" for condition in where_conditions)
    min_total = float(min_total_amount)

    summary = run_query(
        table,
        f"""
        select
            coalesce(cast(DEC_REF_YER as varchar),'Unknown') as DEC_REF_YER,
            coalesce(cast(IDE_CUO_COD as varchar),'Unknown') as IDE_CUO_COD,
            concat(
                coalesce(cast(DEC_REF_YER as varchar),'Unknown'),
                '-',
                coalesce(cast(IDE_CUO_COD as varchar),'Unknown')
            ) as customs_composite_key,
            coalesce(cast(RMT_BNM as varchar),'Unknown') as RMT_BNM,
            coalesce(cast(RMT_BCN as varchar),'Unknown') as exporter_country,
            coalesce(cast(DEC_NAM as varchar),'Unknown') as declarant_name,
            count(*) as remittance_count,
            count(distinct RMT_REF) as ref_count,
            count(distinct RMT_CNM) as remitter_count,
            count(distinct RMT_CUR) as currency_count,
            count(distinct CMP_CON_COD) as consignee_count,
            count(distinct DEC_COD) as declarant_count,
            count(*) filter (where not {NO_CUSDEC_CONDITION}) as import_linked_records,
            count(distinct IDE_REG_NBR) filter (where not {NO_CUSDEC_CONDITION}) as import_declaration_count,
            count(*) filter (where {NO_CUSDEC_CONDITION}) as no_import_records,
            sum(coalesce(RMT_AMT,0)) as total_sent_amount,
            sum(coalesce(RMT_AMT,0)) filter (where not {NO_CUSDEC_CONDITION}) as import_linked_amount,
            sum(coalesce(RMT_AMT,0)) filter (where {NO_CUSDEC_CONDITION}) as no_import_amount,
            sum(coalesce(RMT_OUT,0)) as total_out,
            sum(coalesce(RMT_BAL,0)) as total_balance,
            count(*) filter (where abs(coalesce(RMT_BAL,0))>0.01) as open_balance_transfers,
            case
                when count(*)=0 then null
                else (count(*) filter (where not {NO_CUSDEC_CONDITION})::double/count(*))*100
            end as import_linked_rate,
            case
                when count(*)=0 then null
                else (count(*) filter (where {NO_CUSDEC_CONDITION})::double/count(*))*100
            end as no_import_rate,
            case
                when count(*)=0 then null
                else (count(*) filter (where abs(coalesce(RMT_BAL,0))>0.01)::double/count(*))*100
            end as open_balance_rate,
            case
                when nullif(sum(coalesce(RMT_AMT,0)),0) is null then null
                else (sum(coalesce(RMT_BAL,0))/nullif(sum(coalesce(RMT_AMT,0)),0))*100
            end as balance_to_amount_ratio,
            min(RMT_DAT) as first_rmt_date,
            max(RMT_DAT) as last_rmt_date,
            date_diff('day',min(RMT_DAT),max(RMT_DAT)) as active_days
        from remittances
        {where_clause}
        group by 1,2,3,4,5,6
        having sum(coalesce(RMT_AMT,0)) >= {min_total}
        order by total_sent_amount desc, import_linked_amount desc nulls last, remittance_count desc
        limit {int(row_limit)}
        """,
    ).to_pandas()

    if summary.empty:
        st.info("No BNM vs imports records found for the current filters.")
        return

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Groups", f"{len(summary):,}")
    m2.metric("Sent amount", format_money(summary["total_sent_amount"].fillna(0).sum()))
    m3.metric("Import-linked amount", format_money(summary["import_linked_amount"].fillna(0).sum()))
    m4.metric("No-import amount", format_money(summary["no_import_amount"].fillna(0).sum()))
    m5.metric("Open balance", format_money(summary["total_balance"].fillna(0).sum()))

    fig = px.bar(
        summary.head(30),
        x="total_sent_amount",
        y="RMT_BNM",
        color="customs_composite_key",
        orientation="h",
        title="Top BNM by remittance amount and import key",
        hover_data={
            "DEC_REF_YER": True,
            "IDE_CUO_COD": True,
            "import_declaration_count": True,
            "import_linked_amount": ":,.2f",
            "no_import_amount": ":,.2f",
            "total_balance": ":,.2f",
        },
    )
    style_chart(fig)
    st.plotly_chart(fig, use_container_width=True)

    render_table(summary)
    render_download(
        "Download BNM vs imports summary",
        summary,
        "bnm_vs_imports.csv",
    )


def show_risk_pattern_indicators(table: pa.Table)->None:
    st.subheader("Risk pattern indicators")
    st.caption(
        "These are review indicators only. Use them to prioritize records for document checks, "
        "not as standalone conclusions."
    )

    c1, c2, c3 = st.columns((1.7, 0.8, 0.7))
    search = c1.text_input(
        "Search risk patterns",
        placeholder="CNM, BNM, reference, country, consignee, declarant, or customs key",
        key="risk_pattern_search",
    )
    min_total_amount = c2.number_input(
        "Min total amount",
        min_value=0.0,
        value=0.0,
        step=10000.0,
        key="risk_pattern_min_total",
    )
    row_limit = c3.slider(
        "Rows",
        50,
        5000,
        1000,
        50,
        key="risk_pattern_rows",
    )

    c4, _ = st.columns((0.8, 2.4))
    high_value_threshold = c4.number_input(
        "High value threshold",
        min_value=0.0,
        value=100000.0,
        step=10000.0,
        key="risk_pattern_high_value",
    )

    search_fields = [
        "RMT_CNM",
        "RMT_BNM",
        "RMT_REF",
        "RMT_DES",
        "RMT_BCN",
        "CMP_CON_COD",
        "CMP_CON_NAM",
        "DEC_COD",
        "DEC_NAM",
        "DEC_REF_YER",
        "IDE_CUO_COD",
    ]
    search_condition = ""
    if search:
        joined_fields = ", ".join(f"cast({field} as varchar)" for field in search_fields)
        search_condition = f"lower(concat_ws(' ', {joined_fields})) like {sql_like(search)}"

    search_where = f"where {search_condition}" if search_condition else ""
    search_and = f"and {search_condition}" if search_condition else ""
    min_total = float(min_total_amount)
    high_value = float(high_value_threshold)
    limit = int(row_limit)

    metrics = run_query(
        table,
        f"""
        select
            count(*) as records,
            sum(coalesce(RMT_AMT,0)) as total_amount,
            count(*) filter (where {NO_CUSDEC_CONDITION}) as no_cusdec_records,
            sum(coalesce(RMT_AMT,0)) filter (where {NO_CUSDEC_CONDITION}) as no_cusdec_amount,
            count(*) filter (where abs(coalesce(RMT_BAL,0))>0.01) as open_balance_transfers,
            sum(coalesce(RMT_BAL,0)) as total_open_balance,
            count(distinct RMT_BNM) as bnm_count,
            count(distinct RMT_CNM) as cnm_count
        from remittances
        {search_where}
        """,
    ).row(0, named=True)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Records reviewed", f"{metrics['records'] or 0:,}")
    m2.metric("No CusDec records", f"{metrics['no_cusdec_records'] or 0:,}")
    m3.metric("No CusDec amount", format_money(metrics["no_cusdec_amount"]))
    m4.metric("Open balance", format_money(metrics["total_open_balance"]))

    def show_pattern_result(data, file_name: str, empty_message: str) -> None:
        if data.empty:
            st.info(empty_message)
            return
        render_table(data, height=460)
        render_download("Download pattern results", data, file_name)

    generic_description_condition = """
    (
        RMT_DES is null
        or length(trim(cast(RMT_DES as varchar))) <= 3
        or lower(trim(cast(RMT_DES as varchar))) in ('0', '00', '000', 'goods', 'items', 'material', 'materials', 'sample')
    )
    """

    tabs = st.tabs(
        [
            "BNM many CNM no imports",
            "Repeated pairs",
            "Same-day bursts",
            "CNM many BNM no imports",
            "High value no import",
            "Name similarity",
            "Generic descriptions",
            "Declarant/customs concentration",
            "Open balance ratio",
        ]
    )

    with tabs[0]:
        data = run_query(
            table,
            f"""
            select
                coalesce(cast(RMT_BNM as varchar),'Unknown') as RMT_BNM,
                coalesce(cast(RMT_BCN as varchar),'Unknown') as exporter_country,
                count(*) as no_cusdec_records,
                count(distinct RMT_CNM) as cnm_count,
                count(distinct RMT_REF) as ref_count,
                sum(coalesce(RMT_AMT,0)) as no_cusdec_amount,
                sum(coalesce(RMT_BAL,0)) as total_open_balance,
                max(RMT_AMT) as max_amount,
                min(RMT_DAT) as first_rmt_date,
                max(RMT_DAT) as last_rmt_date,
                date_diff('day',min(RMT_DAT),max(RMT_DAT)) as active_days,
                'Same BNM receiving remittances from multiple CNM with no import link' as risk_indicator,
                'Review possible third-party collection, informal settlement, or missing import write-off.' as review_angle
            from remittances
            where {NO_CUSDEC_CONDITION}
            {search_and}
            group by 1,2
            having count(distinct RMT_CNM) >= 2
               and sum(coalesce(RMT_AMT,0)) >= {min_total}
            order by cnm_count desc, no_cusdec_amount desc
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_bnm_many_cnm_no_imports.csv", "No same-BNM/multiple-CNM no-import pattern found.")

    with tabs[1]:
        data = run_query(
            table,
            f"""
            select
                coalesce(cast(RMT_CNM as varchar),'Unknown') as RMT_CNM,
                coalesce(cast(RMT_BNM as varchar),'Unknown') as RMT_BNM,
                count(*) as transfers,
                count(distinct RMT_REF) as ref_count,
                sum(coalesce(RMT_AMT,0)) as total_amount,
                sum(coalesce(RMT_OUT,0)) as total_out,
                sum(coalesce(RMT_BAL,0)) as total_balance,
                max(RMT_BAL) as max_balance,
                count(*) filter (where abs(coalesce(RMT_BAL,0))>0.01) as open_balance_transfers,
                case
                    when count(*)=0 then null
                    else (count(*) filter (where abs(coalesce(RMT_BAL,0))>0.01)::double/count(*))*100
                end as open_balance_rate,
                case
                    when nullif(sum(coalesce(RMT_AMT,0)),0) is null then null
                    else (sum(coalesce(RMT_BAL,0))/nullif(sum(coalesce(RMT_AMT,0)),0))*100
                end as balance_to_amount_ratio,
                min(RMT_DAT) as first_rmt_date,
                max(RMT_DAT) as last_rmt_date,
                date_diff('day',min(RMT_DAT),max(RMT_DAT)) as active_days,
                'Repeated BNM-CNM pair with remaining balance' as risk_indicator,
                'Review repeated supplier/importer pair for split remittance, unmatched imports, or write-off gaps.' as review_angle
            from remittances
            where 1=1
            {search_and}
            group by 1,2
            having count(*) >= 2
               and abs(sum(coalesce(RMT_BAL,0))) > 0.01
               and sum(coalesce(RMT_AMT,0)) >= {min_total}
            order by abs(total_balance) desc, transfers desc
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_repeated_bnm_cnm_pairs.csv", "No repeated open-balance BNM-CNM pair found.")

    with tabs[2]:
        data = run_query(
            table,
            f"""
            select
                RMT_DAT,
                coalesce(cast(RMT_BNM as varchar),'Unknown') as RMT_BNM,
                coalesce(cast(RMT_BCN as varchar),'Unknown') as exporter_country,
                count(*) as same_day_transfer_count,
                count(distinct RMT_CNM) as cnm_count,
                count(distinct RMT_REF) as ref_count,
                sum(coalesce(RMT_AMT,0)) as same_day_amount,
                sum(coalesce(RMT_BAL,0)) as total_balance,
                max(RMT_AMT) as max_amount,
                'Multiple same-day remittances to same BNM' as risk_indicator,
                'Review for split payments, pooled settlement, or repeated same-day supplier funding.' as review_angle
            from remittances
            where RMT_DAT is not null
            {search_and}
            group by 1,2,3
            having count(*) >= 3
               and sum(coalesce(RMT_AMT,0)) >= {min_total}
            order by same_day_transfer_count desc, same_day_amount desc
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_same_day_bnm_bursts.csv", "No same-day BNM burst pattern found.")

    with tabs[3]:
        data = run_query(
            table,
            f"""
            select
                coalesce(cast(RMT_CNM as varchar),'Unknown') as RMT_CNM,
                count(*) as no_cusdec_records,
                count(distinct RMT_BNM) as bnm_count,
                count(distinct RMT_REF) as ref_count,
                sum(coalesce(RMT_AMT,0)) as no_cusdec_amount,
                sum(coalesce(RMT_BAL,0)) as total_open_balance,
                max(RMT_AMT) as max_amount,
                min(RMT_DAT) as first_rmt_date,
                max(RMT_DAT) as last_rmt_date,
                date_diff('day',min(RMT_DAT),max(RMT_DAT)) as active_days,
                'Same CNM remitting to multiple BNM with no import link' as risk_indicator,
                'Review importer-level non-write-off behavior, informal imports, or third-party supplier routing.' as review_angle
            from remittances
            where {NO_CUSDEC_CONDITION}
            {search_and}
            group by 1
            having count(distinct RMT_BNM) >= 2
               and sum(coalesce(RMT_AMT,0)) >= {min_total}
            order by bnm_count desc, no_cusdec_amount desc
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_cnm_many_bnm_no_imports.csv", "No same-CNM/multiple-BNM no-import pattern found.")

    with tabs[4]:
        data = run_query(
            table,
            f"""
            select
                RMT_DAT,
                RMT_CNM,
                RMT_BNM,
                RMT_BCN,
                RMT_CUR,
                RMT_AMT,
                RMT_BAL,
                RMT_OUT,
                RMT_REF,
                RMT_DES,
                IDE_REG_NBR,
                IDE_REG_DAT,
                date_diff('day',RMT_DAT,current_date) as days_unwritten,
                'High value remittance with no import link' as risk_indicator,
                'Review documents and import write-off before treating as regular settlement.' as review_angle
            from remittances
            where {NO_CUSDEC_CONDITION}
              and coalesce(RMT_AMT,0) >= {high_value}
            {search_and}
            order by RMT_AMT desc, days_unwritten desc nulls last
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_high_value_no_import.csv", "No high-value no-import remittances found.")

    with tabs[5]:
        data = run_query(
            table,
            f"""
            with base as (
                select
                    *,
                    {normalize_name_sql('RMT_BNM')} as normalized_bnm,
                    {normalize_name_sql('RMT_CNM')} as normalized_cnm
                from remittances
                where {NO_CUSDEC_CONDITION}
                {search_and}
            )
            select
                RMT_DAT,
                RMT_CNM,
                RMT_BNM,
                normalized_cnm,
                normalized_bnm,
                case
                    when normalized_bnm = normalized_cnm and normalized_bnm<>'' then 100
                    when normalized_bnm like '%' || normalized_cnm || '%'
                         or normalized_cnm like '%' || normalized_bnm || '%' then 80
                    when levenshtein(normalized_bnm, normalized_cnm) <= 2 then 60
                    else 0
                end as match_score,
                case
                    when normalized_bnm = normalized_cnm and normalized_bnm<>'' then 'Exact Match'
                    when normalized_bnm like '%' || normalized_cnm || '%'
                         or normalized_cnm like '%' || normalized_bnm || '%' then 'Partial Match'
                    when levenshtein(normalized_bnm, normalized_cnm) <= 2 then 'Spelling Variation'
                    else 'Low Similarity'
                end as match_type,
                RMT_AMT,
                RMT_BAL,
                RMT_REF,
                date_diff('day',RMT_DAT,current_date) as days_unwritten,
                'CNM and BNM names are similar on no-import remittance' as risk_indicator,
                'Review for related-party, self-remittance, or circular movement indicators.' as review_angle
            from base
            where normalized_bnm <> ''
              and normalized_cnm <> ''
              and (
                    normalized_bnm = normalized_cnm
                    or normalized_bnm like '%' || normalized_cnm || '%'
                    or normalized_cnm like '%' || normalized_bnm || '%'
                    or levenshtein(normalized_bnm, normalized_cnm) <= 2
              )
              and coalesce(RMT_AMT,0) >= {min_total}
            order by match_score desc, RMT_AMT desc
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_cnm_bnm_name_similarity.csv", "No similar CNM/BNM no-import names found.")

    with tabs[6]:
        data = run_query(
            table,
            f"""
            select
                coalesce(cast(RMT_DES as varchar),'Blank') as RMT_DES,
                count(*) as generic_description_records,
                count(distinct RMT_CNM) as cnm_count,
                count(distinct RMT_BNM) as bnm_count,
                count(*) filter (where {NO_CUSDEC_CONDITION}) as no_cusdec_records,
                sum(coalesce(RMT_AMT,0)) as generic_description_amount,
                sum(coalesce(RMT_BAL,0)) as total_balance,
                min(RMT_DAT) as first_rmt_date,
                max(RMT_DAT) as last_rmt_date,
                'Blank or generic goods description' as risk_indicator,
                'Review weak transaction purpose, missing detail, or unsupported import linkage.' as review_angle
            from remittances
            where {generic_description_condition}
            {search_and}
            group by 1
            having sum(coalesce(RMT_AMT,0)) >= {min_total}
            order by generic_description_amount desc, generic_description_records desc
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_generic_descriptions.csv", "No generic-description pattern found.")

    with tabs[7]:
        data = run_query(
            table,
            f"""
            select
                coalesce(cast(CMP_CON_COD as varchar),'Unknown') as consignee_code,
                coalesce(cast(CMP_CON_NAM as varchar),'Unknown') as consignee_name,
                coalesce(cast(DEC_COD as varchar),'Unknown') as declarant_code,
                coalesce(cast(DEC_NAM as varchar),'Unknown') as declarant_name,
                concat(
                    coalesce(cast(DEC_REF_YER as varchar),'Unknown'),
                    '-',
                    coalesce(cast(IDE_CUO_COD as varchar),'Unknown')
                ) as customs_composite_key,
                count(*) as remittance_count,
                count(distinct RMT_REF) as ref_count,
                count(distinct RMT_CNM) as remitter_count,
                count(distinct RMT_BNM) as beneficiary_count,
                count(*) filter (where {NO_CUSDEC_CONDITION}) as no_cusdec_records,
                sum(coalesce(RMT_AMT,0)) as total_sent_amount,
                sum(coalesce(RMT_BAL,0)) as total_balance,
                sum(coalesce(RMT_OUT,0)) as total_out,
                min(RMT_DAT) as first_rmt_date,
                max(RMT_DAT) as last_rmt_date,
                date_diff('day',min(RMT_DAT),max(RMT_DAT)) as active_days,
                'Concentrated consignee-declarant-customs key remittance amount' as risk_indicator,
                'Review concentrated declaration channel and repeated consignee/declarant relationship.' as review_angle
            from remittances
            where 1=1
            {search_and}
            group by 1,2,3,4,5
            having sum(coalesce(RMT_AMT,0)) >= {min_total}
            order by total_sent_amount desc, remittance_count desc
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_declarant_customs_concentration.csv", "No declarant/customs-key concentration found.")

    with tabs[8]:
        data = run_query(
            table,
            f"""
            select
                coalesce(cast(RMT_CNM as varchar),'Unknown') as RMT_CNM,
                coalesce(cast(RMT_BNM as varchar),'Unknown') as RMT_BNM,
                count(*) as transfers,
                count(distinct RMT_REF) as ref_count,
                sum(coalesce(RMT_AMT,0)) as total_amount,
                sum(coalesce(RMT_BAL,0)) as total_balance,
                max(RMT_BAL) as max_balance,
                avg(coalesce(RMT_BAL,0)) as avg_balance,
                count(*) filter (where abs(coalesce(RMT_BAL,0))>0.01) as open_balance_transfers,
                case
                    when count(*)=0 then null
                    else (count(*) filter (where abs(coalesce(RMT_BAL,0))>0.01)::double/count(*))*100
                end as open_balance_rate,
                case
                    when nullif(sum(coalesce(RMT_AMT,0)),0) is null then null
                    else (sum(coalesce(RMT_BAL,0))/nullif(sum(coalesce(RMT_AMT,0)),0))*100
                end as balance_to_amount_ratio,
                min(RMT_DAT) as first_rmt_date,
                max(RMT_DAT) as last_rmt_date,
                date_diff('day',min(RMT_DAT),max(RMT_DAT)) as active_days,
                'High remaining balance versus sent amount' as risk_indicator,
                'Review remittance write-off, unmatched imports, or excess remittance balances.' as review_angle
            from remittances
            where 1=1
            {search_and}
            group by 1,2
            having sum(coalesce(RMT_AMT,0)) >= {min_total}
               and abs(sum(coalesce(RMT_BAL,0))) > 0.01
               and abs(sum(coalesce(RMT_BAL,0)) / nullif(sum(coalesce(RMT_AMT,0)),0)) >= 0.20
            order by abs(balance_to_amount_ratio) desc, abs(total_balance) desc
            limit {limit}
            """,
        ).to_pandas()
        show_pattern_result(data, "risk_open_balance_ratio.csv", "No high open-balance-ratio pattern found.")



# ---------------------------------------------------------------------------
# Main app
# ---------------------------------------------------------------------------

def main() -> None:
    st.title("Remittance Review Dashboard")

    with st.sidebar:
        st.header("Data source")
        uploaded = st.file_uploader("Upload Excel workbook", type=["xlsx", "xlsm"])
        if uploaded:
            with NamedTemporaryFile(delete=False, suffix=Path(uploaded.name).suffix) as tmp:
                tmp.write(uploaded.getbuffer())
                source_path = Path(tmp.name)
        else:
            source_path = DEFAULT_EXCEL_PATH
        st.caption(str(source_path))

    if not source_path.exists():
        st.error(f"Excel file not found: {source_path}")
        st.stop()

    sheets = get_sheet_names(str(source_path))
    preferred_sheet = DEFAULT_SHEET if DEFAULT_SHEET in sheets else sheets[0]
    with st.sidebar:
        sheet_name = st.selectbox("Sheet", sheets, index=sheets.index(preferred_sheet))

    df = load_excel(str(source_path), sheet_name)
    filtered = apply_sidebar_filters(df)
    table = arrow_table(filtered)

    st.caption(
        f"Loaded {df.height:,} rows and {df.width:,} columns from `{sheet_name}`. "
        f"Current filters show {filtered.height:,} rows."
    )

    report = st.radio(
        "Report",
        [
            "Balance remittances",
            "Balance by exporter (BNM)",
            "No CusDec / not written off",
            "Consignee-declarant mapping",
            "BNM-CNM mapping",
            "Consignee-declarant amount",
            "BNM vs imports",
            "Risk pattern indicators",
            "Data quality",
        ],
        horizontal=True,
    )

    if report == "Balance remittances":
        show_balance_remittances(table)
    elif report == "Balance by exporter (BNM)":
        show_balance_by_exporter(table)
    elif report == "No CusDec / not written off":
        show_no_cusdec_report(table)
    elif report == "Consignee-declarant mapping":
        show_consignee_declarant_mapping(table)
    elif report == "BNM-CNM mapping":
        show_bnm_cnm_mapping(table)
    elif report == "Consignee-declarant amount":
        show_consignee_declarant_remittance_amount(table)
    elif report == "BNM vs imports":
        show_bnm_vs_imports(table)
    elif report == "Risk pattern indicators":
        show_risk_pattern_indicators(table)
    else:
        show_data_quality(table)


if __name__ == "__main__":
    main()
